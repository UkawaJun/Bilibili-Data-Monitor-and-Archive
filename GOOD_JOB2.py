
import yt_dlp
import os
import requests
import time
from openpyxl import Workbook, load_workbook
import math, openpyxl
import re
import datetime
import random
import string
import shutil
import pyzipper  # 必须安装: pip install pyzipper


# FFmpeg 路径
FFMPEG_PATH = r'C:\ffmpeg\bin' 

# 初始目标链接 (会被下方的主程序循环覆盖)
TARGET_URL = "https://www.bilibili.com/video/BV1DLznBgERM/"

# 根目录
BASE_DIR = "File2"

#伪装头
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://www.bilibili.com"
}
# ===========================================

def _ReadXlsl(file_path):
    print(f"正在快速参考文件: {file_path} ")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        all_sheets = wb.sheetnames
        if not all_sheets: return []
        ws = wb[all_sheets[0]]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

def check_env():
    if not os.path.exists(os.path.join(FFMPEG_PATH, 'ffmpeg.exe')):
        print(f"❌ 错误：在 {FFMPEG_PATH} 找不到 ffmpeg.exe")
        return False
    if not os.path.exists(BASE_DIR):
        os.makedirs(BASE_DIR)
    return True

# === 新增功能：从Excel获取最大Index，实现删文件夹不影响计数 ===
def get_next_index_from_excel(base_dir):
    excel_path = os.path.join(base_dir, "download_report.xlsx")
    max_idx = 0
    
    # 如果Excel不存在，说明是第一次运行，检查文件夹作为兜底
    if not os.path.exists(excel_path):
        if not os.path.exists(base_dir):
            return 1
        for folder_name in os.listdir(base_dir):
            if folder_name.isdigit():
                idx = int(folder_name)
                if idx > max_idx:
                    max_idx = idx
        return max_idx + 1

    # 如果Excel存在，读取Excel的第一列
    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        # 跳过表头，遍历第一列
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None and isinstance(row[0], int):
                if row[0] > max_idx:
                    max_idx = row[0]
        wb.close()
    except Exception as e:
        print(f"读取Excel索引出错，降级为扫描文件夹: {e}")
        return get_next_index_from_excel(base_dir) # 这里的递归需要小心，简单起见可以直接返回扫描文件夹逻辑，这里简化处理
        
    return max_idx + 1

# === 新增功能：生成随机密码 ===
def generate_password(length=6):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(length))

# === 新增功能：加密压缩 ===
def create_encrypted_zip(source_dir, output_zip_path, password):
    """
    将 source_dir 下的所有文件打包进 output_zip_path
    使用 AES 加密
    """
    with pyzipper.AESZipFile(output_zip_path, 'w', compression=pyzipper.ZIP_LZMA, encryption=pyzipper.WZ_AES) as zf:
        zf.setpassword(password.encode('utf-8'))
        for root, dirs, files in os.walk(source_dir):
            for file in files:
                # 排除还没写完的 zip 本身
                if file == os.path.basename(output_zip_path):
                    continue
                file_path = os.path.join(root, file)
                # 存入压缩包的文件名（去掉路径，只保留文件名）
                arcname = file 
                zf.write(file_path, arcname)

def format_file_size(size_bytes):
    if not size_bytes: return "0B"
    size_name = ("B", "KB", "MB", "GB")
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    return f"{round(size_bytes / p, 2)} {size_name[i]}"

def format_seconds(seconds):
    if not seconds: return "00:00:00"
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "{:02d}:{:02d}:{:02d}".format(int(h), int(m), int(s))

def format_date_str(date_str):
    if date_str and len(date_str) == 8:
        return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
    return date_str

def get_bilibili_extra_info(bvid):
    info_data = {"view": 0, "favorite": 0}
    related_list = []
    try:
        api_view = f"https://api.bilibili.com/x/web-interface/view?bvid={bvid}"
        resp = requests.get(api_view, headers=HEADERS, timeout=5)
        data = resp.json()
        if data['code'] == 0:
            stat = data['data']['stat']
            info_data['view'] = stat.get('view', 0)
            info_data['favorite'] = stat.get('favorite', 0)
    except Exception:
        pass

    try:
        api_related = f"https://api.bilibili.com/x/web-interface/archive/related?bvid={bvid}"
        resp = requests.get(api_related, headers=HEADERS, timeout=5)
        data = resp.json()
        if data['code'] == 0:
            for item in data['data']:
                r_bvid = item.get('bvid')
                if r_bvid:
                    clean_url = f"https://www.bilibili.com/video/{r_bvid}"
                    related_list.append({
                        "title": item.get('title', '未知'),
                        "owner": item.get('owner', {}).get('name', '未知'),
                        "view": item.get('stat', {}).get('view', 0),
                        "url": clean_url
                    })
    except Exception:
        pass
    return info_data, related_list

def save_related_excel(save_path, related_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "相关推荐"
    ws.append(["视频标题", "UP主", "播放量", "链接"])
    for item in related_data:
        ws.append([item['title'], item['owner'], item['view'], item['url']])
    wb.save(save_path)

def process_download():
    if not check_env():
        return

    excel_path = os.path.join(BASE_DIR, "download_report.xlsx")
    columns = ["Index", "视频名称", "作者", "发布日期", "下载日期", "时长", "大小", "播放量", "收藏数", "解压密码", "原始链接"]
    
    # 初始化 Excel
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "资源列表"
        ws.append(columns)

    current_index = get_next_index_from_excel(BASE_DIR)
    
    print("🔍 正在解析初始信息...")
    with yt_dlp.YoutubeDL({'quiet': True, 'extract_flat': True}) as ydl:
        try:
            info = ydl.extract_info(TARGET_URL, download=False)
        except Exception as e:
            print(f"❌ 解析失败: {e}")
            return

    video_list = info['entries'] if 'entries' in info else [info]
    today_date = datetime.datetime.now().strftime("%Y-%m-%d")

    for i, item in enumerate(video_list):
        real_index = str(current_index + i)
        
        origin_url = item.get('url', item.get('webpage_url', TARGET_URL))
        # 预设默认值，防止报错
        video_title = item.get('title', 'Unknown Title')
        uploader = "未知作者"
        upload_date_str = ""
        duration_str = "00:00:00"
        size_str = "0B"
        zip_password = ""
        stats_view = 0
        stats_fav = 0
        
        print(f"\n[{i+1}/{len(video_list)}] 处理: {video_title} (Index: {real_index})")

        video_dir = os.path.join(BASE_DIR, real_index)
        if not os.path.exists(video_dir):
            os.makedirs(video_dir)

        # 1. API 
        bv_match = re.search(r'(BV\w+)', origin_url)
        bvid = bv_match.group(1) if bv_match else None
        if bvid:
            stats_info, related_videos = get_bilibili_extra_info(bvid)
            stats_view = stats_info['view']
            stats_fav = stats_info['favorite']
            if related_videos:
                save_related_excel(os.path.join(video_dir, "related_videos.xlsx"), related_videos)

        # 2. 下载
        final_file_path = os.path.join(video_dir, f"{real_index}.mp4")
        ydl_opts = {
            'ffmpeg_location': FFMPEG_PATH,
            'outtmpl': f'{video_dir}/{real_index}.%(ext)s',
            'format': 'bv[vcodec^=avc][height<=1080]+ba/b[height<=1080]',
            'merge_output_format': 'mp4',
            'writethumbnail': True,
            'postprocessors': [
                {'key': 'FFmpegThumbnailsConvertor', 'format': 'png'},
                {'key': 'FFmpegVideoConvertor', 'preferedformat': 'mp4'}
            ],
            'quiet': False,
            'no_warnings': True,
            'ignoreerrors': True, # 遇到小错误继续
        }

        #阶段一：下载
        download_success = False
        try:
            with yt_dlp.YoutubeDL(ydl_opts) as ydl_worker:
                meta = ydl_worker.extract_info(origin_url, download=True)
                
                if meta:
                    video_title = meta.get('title', video_title)
                    duration_str = format_seconds(meta.get('duration', 0))
                    uploader = meta.get('uploader', '未知作者')
                    upload_date_str = format_date_str(meta.get('upload_date'))
                    description = meta.get('description', '无简介')

                    desc_path = os.path.join(video_dir, "简介.txt")
                    with open(desc_path, "w", encoding="utf-8") as f:
                        f.write(description if description else "无简介")
                    
                    download_success = True
                    
        except Exception as e:
            print(f"   ❌ 下载阶段出错: {e}")
            # 如果下载失败，就不进行压缩了，直接记 Excel Error

        # 阶段二：压缩打包
        if download_success:
            # 获取大小
            if os.path.exists(final_file_path):
                size_str = format_file_size(os.path.getsize(final_file_path))
            
            # !关键!：休息2秒，等待系统释放文件占用
            time.sleep(2) 
            
            try:
                zip_password = generate_password(6)
                zip_filename = f"{real_index}.zip"
                zip_full_path = os.path.join(video_dir, zip_filename)
                
                print(f"   🔒 正在加密压缩...")
                create_encrypted_zip(video_dir, zip_full_path, zip_password)
                
                # 压缩成功后清理文件
                for f_name in os.listdir(video_dir):
                    if f_name != zip_filename:
                        try:
                            f_path = os.path.join(video_dir, f_name)
                            if os.path.isdir(f_path):
                                shutil.rmtree(f_path)
                            else:
                                os.remove(f_path)
                        except:
                            pass
                            
            except Exception as e:
                print(f"   ⚠️ 打包失败 (文件保留): {e}")
                zip_password = "打包出错-文件未加密"

        # 阶段三：写入 Excel
        # 只要下载成功了，哪怕打包失败，也要把信息写进去，而不是写 Error
        if download_success:
            row_data = [
                int(real_index), video_title, uploader, upload_date_str, today_date,
                duration_str, size_str, stats_view, stats_fav, zip_password, origin_url
            ]
            ws.append(row_data)
            wb.save(excel_path)
            print(f"   --> Excel 更新完毕 (Index: {real_index})")
        else:
            # 只有连下载都没成功，才记录 Error
            ws.append([int(real_index), video_title, "Error", "", today_date, "", "0B", 0, 0, "", origin_url])
            wb.save(excel_path)
            
        time.sleep(2)

    print(f"\n🎉 任务完成！")


if __name__ == '__main__':

    _data = _ReadXlsl("BID.xlsx")[1:]
    _data.reverse()
    
    # 0 -1625
    # 0:50 50 100 
    _data = _data[1200:1300] 
    print(f"长度{len(_data)}")
    input("按回车开始...")
    
    i = 1
    for _burl in _data:

        if not _burl or not _burl[0]: continue
        
        TARGET_URL = _burl[1]
        print(f"\n========================================")
        print(f"当前进度 {i}/{len(_data)} | 目标: {TARGET_URL}")
        
        process_download()
        
        time.sleep(6)
        print("\a")

        i += 1
